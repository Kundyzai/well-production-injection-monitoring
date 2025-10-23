import os
import glob
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.utils import column_index_from_string, get_column_letter
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import time
import sys
import logging
import traceback

# === Настройка путей ===
FOLDER_PATH = r"C:\Your\Project\Path\Injection folder"
INJECTION_FILE = os.path.join(FOLDER_PATH, 'Injection.xlsx')

# === Логирование (только в консоль) ===
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger()

# === Вспомогательные функции ===
def get_next_column(col_letter, offset=1):
    col_idx = column_index_from_string(col_letter)
    return get_column_letter(col_idx + offset)


def copy_formulas(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=source_row, column=col)
        if cell.data_type == 'f':
            formula = cell.value
            for r in range(source_row - 5, source_row + 5):
                if str(r) in formula:
                    formula = formula.replace(str(r), str(target_row - (source_row - r)))
            ws.cell(row=target_row, column=col).value = formula


def process_injection_data(file_path):
    """Обрабатывает новый Excel-файл с данными инъекций."""
    try:
        logger.info(f"Начинаю обработку файла: {os.path.basename(file_path)}")

        # Извлекаем дату из имени файла
        date_str = os.path.basename(file_path).split(" ")[-1].replace(".xlsx", "")
        date = datetime.strptime(date_str, "%d.%m.%Y")

        # Загружаем данные
        df_daily = pd.read_excel(file_path, sheet_name="Daily_Production_EN_CB", header=None)

        # Извлекаем данные по скважинам
        well_data = {}
        processed_wells = set()

        for row in range(159, 169):
            well_name = df_daily.iloc[row, 3]  # D
            if pd.isna(well_name) or not isinstance(well_name, str) or not well_name.startswith("well_"):
                continue
            if well_name in processed_wells:
                continue

            online_hrs = df_daily.iloc[row, 5]  # F
            thp = df_daily.iloc[row, 7]         # H
            water_injection = df_daily.iloc[row, 30]  # AE

            well_data[well_name] = (online_hrs, thp, water_injection)
            processed_wells.add(well_name)

        wb = load_workbook(INJECTION_FILE)
        ws = wb["2025"]

        # Поиск строки с датой
        target_row = None
        for row in range(3, ws.max_row + 1):
            cell_date = ws[f"A{row}"].value
            if isinstance(cell_date, datetime) and cell_date.date() == date.date():
                target_row = row
                break

        # Если дата не найдена — вставляем строку
        if target_row is None:
            last_date_row = 3
            for row in range(3, ws.max_row + 1):
                if isinstance(ws[f"A{row}"].value, datetime):
                    last_date_row = row
            ws.insert_rows(last_date_row + 1)
            target_row = last_date_row + 1
            ws[f"A{target_row}"] = date
            copy_formulas(ws, last_date_row, target_row)
            logger.info(f"Добавлена новая строка для даты {date.date()}")

        # Соответствие скважин и колонок
        well_columns = {
            "well_92": "B", "well_93": "G", "well_94": "L",
            "well_95": "Q", "well_96": "V", "well_97": "AA",
            "well_98": "AF", "well_99": "AK", "well_100": "AP",
            "well_101": "AU"
        }

        # Перенос данных
        for well, (online_hrs, thp, water_injection) in well_data.items():
            if well in well_columns:
                base_col = well_columns[well]
                ws[f"{base_col}{target_row}"] = online_hrs
                ws[f"{get_next_column(base_col)}{target_row}"] = thp
                ws[f"{get_next_column(base_col, 2)}{target_row}"] = water_injection

        wb.save(INJECTION_FILE)
        logger.info("✅ Данные успешно перенесены!")
        logger.info(f"Обработаны скважины: {list(well_data.keys())}")

    except Exception as e:
        logger.error(f"❌ Ошибка при обработке файла: {e}")
        logger.error(traceback.format_exc())


# === Наблюдение за файлами ===
class FileHandler(FileSystemEventHandler):
    def __init__(self):
        self.processed_files = set()

    def on_created(self, event):
        if not event.is_directory and event.src_path.endswith('.xlsx'):
            time.sleep(2)
            filename = os.path.basename(event.src_path)
            if filename.startswith('Daily_Production_EN_CB') and filename not in self.processed_files:
                logger.info(f"📁 Обнаружен новый файл: {filename}")
                self.processed_files.add(filename)
                process_injection_data(event.src_path)


def process_existing_files():
    files = glob.glob(os.path.join(FOLDER_PATH, 'Daily_Production_EN_CB *.xlsx'))
    if files:
        files.sort(key=os.path.getctime, reverse=True)
        latest_file = files[0]
        logger.info(f"⚙️ Обрабатываю последний существующий файл: {os.path.basename(latest_file)}")
        process_injection_data(latest_file)
    else:
        logger.warning("⚠️ Файлы Daily_Production_EN_CB не найдены")


def start_injection_monitoring():
    logger.info("🚀 Запуск мониторинга Injection файлов...")
    process_existing_files()

    event_handler = FileHandler()
    observer = Observer()
    observer.schedule(event_handler, FOLDER_PATH, recursive=False)
    observer.start()

    logger.info(f"👀 Наблюдение запущено за папкой: {FOLDER_PATH}")
    logger.info("Ожидание новых файлов...")

    return observer


def stop_monitoring(observer):
    observer.stop()
    observer.join()
    logger.info("🛑 Наблюдение остановлено")


if __name__ == "__main__":
    if not os.path.exists(FOLDER_PATH):
        logger.error(f"Папка {FOLDER_PATH} не существует!")
        sys.exit(1)

    if not os.path.exists(INJECTION_FILE):
        logger.error(f"Файл {INJECTION_FILE} не существует!")
        sys.exit(1)

    observer = start_injection_monitoring()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        stop_monitoring(observer)
    except Exception as e:
        logger.error(f"❌ Неожиданная ошибка: {e}")
        stop_monitoring(observer)
