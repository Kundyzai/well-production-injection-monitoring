import os
import glob
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import logging
import re
import time
import shutil
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# Путь к директории
PATH = r"C:\Your\Project\Path\CPU_update"
CPU_FILE = os.path.join(PATH, "CPU_update.xlsm")

# --- Проверка имени скважины ---
def is_valid_well(name):
    if not isinstance(name, str):
        return False
    m = re.match(r'^well_(\d+)$', name.strip())
    if m:
        num = int(m.group(1))
        return 1 <= num <= 91
    return False

# --- Извлечение даты из имени файла ---
def extract_date_from_filename(filename):
    m = re.search(r'(\d{2}\.\d{2}\.\d{4})', filename)
    if m:
        return datetime.strptime(m.group(1), '%d.%m.%Y')
    return None

# --- Определение имени листа по месяцу ---
def get_month_sheet_name(date_obj):
    months = {
        1:'january',2:'february',3:'march',4:'april',
        5:'may',6:'june',7:'july',8:'august',
        9:'september',10:'october',11:'november',12:'december'
    }
    return months.get(date_obj.month, 'unknown')

# --- Проверка, открыт ли файл Excel ---
def is_file_locked(filepath):
    try:
        os.rename(filepath, filepath)
        return False
    except PermissionError:
        return True

# --- Безопасное открытие книги ---
def safe_load_workbook(filepath):
    if not os.path.exists(filepath):
        logging.warning(f"Файл {filepath} не найден, будет создан новый.")
        return None

    if is_file_locked(filepath):
        logging.warning(f"Файл {os.path.basename(filepath)} открыт в Excel. Закрой его перед обновлением.")
        return None

    try:
        wb = load_workbook(filepath, keep_vba=True)
        return wb
    except Exception as e:
        logging.error(f"Ошибка при открытии {filepath}: {e}")
        return None

# --- Поиск колонки по дате ---
def find_date_column(ws, target_date):
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if not val:
            continue
        if isinstance(val, datetime):
            if val.date() == target_date.date():
                return col
        elif isinstance(val, str):
            try:
                d = datetime.strptime(val.strip(), '%d.%m.%Y')
                if d.date() == target_date.date():
                    return col
            except ValueError:
                continue
    return None

# --- Создание новой колонки с датой ---
def create_date_column(ws, date_obj):
    last_col = 1
    while ws.cell(row=1, column=last_col).value is not None:
        last_col += 9  # каждая дата занимает 9 колонок

    ws.cell(row=1, column=last_col).value = date_obj
    headers = ['Well_name', 'status', 'RPM', 'Oil, m3', 'Fluid, m3', 'Water, m3', 'Gas, m3', 'GOR, m3/m3', 'WC, %']
    for i, header in enumerate(headers):
        ws.cell(row=2, column=last_col + i).value = header

    return last_col

# --- Безопасное сохранение книги (одна резервная копия) ---
def safe_save_workbook(wb, filepath):
    try:
        backup_file = filepath.replace(".xlsm", "_backup.xlsm")

        # Создание или обновление резервной копии
        if os.path.exists(filepath):
            shutil.copy2(filepath, backup_file)
            logging.info(f"Резервная копия обновлена: {os.path.basename(backup_file)}")
        else:
            logging.info("Исходный файл не найден, резервная копия не создана.")

        # Сохранение во временный файл
        temp_file = filepath + ".tmp"
        wb.save(temp_file)
        os.replace(temp_file, filepath)

        logging.info(f"Изменения успешно сохранены в {os.path.basename(filepath)}")
    except Exception as e:
        logging.error(f"Ошибка при сохранении файла: {e}")

# --- Основное обновление данных ---
def update_cpu_data(source_file):
    logging.info(f"Обработка файла: {os.path.basename(source_file)}")

    date_obj = extract_date_from_filename(os.path.basename(source_file))
    if not date_obj:
        logging.error("Не удалось определить дату из имени файла.")
        return

    wb = safe_load_workbook(CPU_FILE)
    if wb is None:
        return  # если открыть не удалось — выходим

    sheet_name = get_month_sheet_name(date_obj)
    if sheet_name not in wb.sheetnames:
        logging.info(f"Создаю новый лист: {sheet_name}")
        ws = wb.create_sheet(sheet_name)
        col = create_date_column(ws, date_obj)
    else:
        ws = wb[sheet_name]
        col = find_date_column(ws, date_obj)
        if not col:
            logging.info(f"Создаю новую колонку для даты {date_obj.strftime('%d.%m.%Y')}")
            col = create_date_column(ws, date_obj)

    df = pd.read_excel(source_file, sheet_name='CPU_Production_EN', header=7, usecols='N,P,R,S,T', nrows=300)
    df.columns = ['Well_name', 'Oil', 'Water', 'Gas', 'RPM']
    df = df[df['Well_name'].apply(is_valid_well)]
    logging.info(f"Найдено {len(df)} скважин в исходном файле")

    data_dict = {}
    for _, row in df.iterrows():
        data_dict[row['Well_name']] = {
            'RPM': row['RPM'],
            'Oil': row['Oil'],
            'Water': row['Water'],
            'Gas': row['Gas'],
            'Fluid': (row['Oil'] or 0) + (row['Water'] or 0)
        }

    updated = 0
    for r in range(3, ws.max_row + 1):
        well = ws.cell(row=r, column=col).value
        if well in data_dict:
            d = data_dict[well]
            ws.cell(row=r, column=col + 2).value = d['RPM']
            ws.cell(row=r, column=col + 3).value = d['Oil']
            ws.cell(row=r, column=col + 4).value = d['Fluid']
            ws.cell(row=r, column=col + 5).value = d['Water']
            ws.cell(row=r, column=col + 6).value = d['Gas']
            ws.cell(row=r, column=col + 7).value = f"=IF({ws.cell(row=r, column=col+6).coordinate}=0,0,{ws.cell(row=r, column=col+6).coordinate}/{ws.cell(row=r, column=col+3).coordinate})"
            ws.cell(row=r, column=col + 8).value = f"=IF({ws.cell(row=r, column=col+5).coordinate}=0,0,{ws.cell(row=r, column=col+5).coordinate}/{ws.cell(row=r, column=col+4).coordinate}*100)"
            updated += 1
            del data_dict[well]

    for well_name, data in data_dict.items():
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=col).value = well_name
        ws.cell(row=new_row, column=col + 2).value = data['RPM']
        ws.cell(row=new_row, column=col + 3).value = data['Oil']
        ws.cell(row=new_row, column=col + 4).value = data['Fluid']
        ws.cell(row=new_row, column=col + 5).value = data['Water']
        ws.cell(row=new_row, column=col + 6).value = data['Gas']
        ws.cell(row=new_row, column=col + 7).value = f"=IF({ws.cell(row=new_row, column=col+6).coordinate}=0,0,{ws.cell(row=new_row, column=col+6).coordinate}/{ws.cell(row=new_row, column=col+3).coordinate})"
        ws.cell(row=new_row, column=col + 8).value = f"=IF({ws.cell(row=new_row, column=col+5).coordinate}=0,0,{ws.cell(row=new_row, column=col+5).coordinate}/{ws.cell(row=new_row, column=col+4).coordinate}*100)"
        updated += 1

    safe_save_workbook(wb, CPU_FILE)
    logging.info(f"✅ Обновлено {updated} скважин на листе {sheet_name} для {date_obj.strftime('%d.%m.%Y')}.")

# --- Класс для отслеживания новых файлов ---
class NewFileHandler(FileSystemEventHandler):
    def __init__(self):
        self.processed_files = set()

    def on_created(self, event):
        if not event.is_directory and event.src_path.endswith('.xlsx'):
            time.sleep(2)
            filename = os.path.basename(event.src_path)
            if filename.startswith('CPU_Production_EN') and filename not in self.processed_files:
                logging.info(f"Обнаружен новый файл: {filename}")
                self.processed_files.add(filename)
                update_cpu_data(event.src_path)

# --- Основной цикл мониторинга ---
def start_monitoring():
    logging.info("Запуск мониторинга папки...")
    files = glob.glob(os.path.join(PATH, "CPU_Production_EN *.xlsx"))
    if files:
        latest = max(files, key=os.path.getctime)
        logging.info(f"Обрабатываю последний файл: {os.path.basename(latest)}")
        update_cpu_data(latest)

    event_handler = NewFileHandler()
    observer = Observer()
    observer.schedule(event_handler, PATH, recursive=False)
    observer.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

if __name__ == "__main__":
    start_monitoring()